import openpyxl


class ExcelHandler(object):

    def __init__(self, url):
        self.url = url
        self.wb = openpyxl.load_workbook(self.url)

    def get_workbook(self):
        return self.wb

    def get_worksheet(self, sheet_name):
        worksheet = self.get_workbook()[sheet_name]
        return worksheet

    def get_cell(self, sheet_name, row, column):
        cell = self.get_worksheet(sheet_name).cell(row, column)
        return cell

    # def get_init_data(self):
    #     cell = self.get_cell('init', 1, 2)
    #     mobile = cell.value
    #     return mobile

    # def update_init_data(self, value):
    #     wb = self.get_workbook()
    #     sheet = wb['init']
    #     sheet.cell(1, 2, value=value)
    #     wb.save(self.url)

    # @staticmethod
    # def get_num(num):
    #     num += 1
    #     while True:
    #         conn = pymysql.connect(
    #             host="120.78.128.25",
    #             port=3306,
    #             user="future",
    #             password="123456",
    #             charset='utf8',
    #             cursorclass=DictCursor
    #         )
    #         cursor = conn.cursor()
    #         cursor.execute("select * from futureloan.member where mobile_phone={};".format(num))
    #         sql = cursor.fetchone()
    #         cursor.close()
    #         conn.close()
    #         if not sql:
    #             return num
    #         else:
    #             num += 1

    def data_trans(self, sheet_name):
        cases = list(self.get_worksheet(sheet_name).values)
        # mobile = self.get_init_data()
        data = []
        for case in cases[1:]:
            d = {}
            for i, v in enumerate(case):
                # if isinstance(case[i], str) and case[i].find("${mobile}") != -1:
                #     v = case[i].replace("${mobile}", str(mobile))
                d[cases[0][i]] = v
            data.append(d)
        # self.update_init_data(self.get_num(mobile))
        self.close_wb()
        return data

    def data_write(self, sheet_name, row, column, new_data):
        cell = self.get_cell(sheet_name, row, column)
        cell.value = new_data
        self.save_wb()
        self.close_wb()

    def save_wb(self):
        self.wb.save(self.url)

    def close_wb(self):
        self.wb.close()


if __name__ == '__main__':
    url = r'E:\python_test\test_demo\data\cases.xlsx'
    do_excel = ExcelHandler(url)
    my_data = do_excel.data_trans('Sheet1')
    print(my_data)

